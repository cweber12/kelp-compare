"""Adapter-level tests for `hobo_xlsx` (docs/06).

Two kinds of test here. Against the reference files in `tests/fixtures/`, the
adapter must reproduce the docs/06 findings that `test_fixtures_hobo.py` pins.
Against small synthetic workbooks built in `tmp_path`, it must handle the cases
the reference files cannot show us -- an unregistered serial, a different unit
and timezone, a second series, a gap in the cadence.

No network: every input is a local file, and the synthetic ones are written
here rather than committed, so the two reference binaries stay the only fixtures
of record.
"""

from __future__ import annotations

import json
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from kelpcompare.adapters import hobo_xlsx
from kelpcompare.adapters.base import REGISTRY_GATE
from kelpcompare.registry import Deployment, find_deployment, load_registry

REPO_ROOT = Path(__file__).resolve().parents[1]
FIX = Path(__file__).parent / "fixtures"
ORIGINAL = FIX / "Tidbit_1__22506632__2026-08-01_07_44_27_PDT__Data_PDT_.xlsx"
EDITED = FIX / "yellow_buoy_temps.xlsx"
REGISTRY_PATH = REPO_ROOT / "data" / "registry" / "sites.json"

DEGREE = "°"  # the fixtures use U+00B0, not a lookalike
SERIES_COLUMN = f"Tidbit 1 , {DEGREE}F"
KNOWN_SERIAL = "22506632"


def local(*parts: int) -> datetime:
    """A naive local timestamp -- exactly what a HOBO file carries.

    The adapter must not attach a timezone (docs/06 s4: the header token is
    reported, applying it is the normalizer's job), so these fixtures and
    assertions are naive on purpose.
    """
    return datetime(*parts)  # noqa: DTZ001 -- naive by design, see above


#: Inside the reviewed deployment, so synthetic files land in PDT like the fixtures.
DEFAULT_START = local(2026, 7, 11, 7, 0)


@pytest.fixture(scope="module")
def registry():
    return load_registry(REGISTRY_PATH)


# --------------------------------------------------------------------------
# Synthetic workbook builder
# --------------------------------------------------------------------------


def write_hobo_xlsx(
    path: Path,
    *,
    serial: str = KNOWN_SERIAL,
    tz: str = "PDT",
    series: tuple[tuple[str, str, list[float]], ...] = (
        ("Tidbit 1", f"{DEGREE}F", [60.0, 61.0, 62.0]),
    ),
    start: datetime = DEFAULT_START,
    interval_minutes: int = 10,
    gap_after: int | None = None,
    product: str = "MX2204",
    deployment_number: int = 3,
    reported_samples: int | None = None,
    sheets: tuple[str, ...] = ("Data", "Events", "Details"),
) -> Path:
    """Write a minimal but structurally faithful HOBOconnect workbook.

    Defaults produce a clean *original*: contiguous `#`, no extra columns, no
    formulas, and `Details` statistics that agree with the data.
    """
    book = openpyxl.Workbook()
    book.remove(book.active)

    n = len(series[0][2])
    stamps = []
    moment = start
    for index in range(n):
        stamps.append(moment)
        step = interval_minutes * (2 if gap_after is not None and index == gap_after else 1)
        moment = moment + timedelta(minutes=step)

    if "Data" in sheets:
        sheet = book.create_sheet("Data")
        sheet.append(["#", f"Date-Time ({tz})"] + [f"{name} , {unit}" for name, unit, _ in series])
        for index, stamp in enumerate(stamps):
            sheet.append([index + 1, stamp] + [values[index] for _, _, values in series])

    if "Events" in sheets:
        sheet = book.create_sheet("Events")
        sheet.append(["#", f"Date-Time ({tz})", "Host Connected", "End of File", "Started"])
        sheet.append([1, stamps[0], None, None, "Logged"])
        sheet.append([2, stamps[-1], None, "Logged", None])

    if "Details" in sheets:
        sheet = book.create_sheet("Details")
        for row in (
            ["Details", None, None, None],
            ["Devices", None, None, None],
            [None, "Device Info", None, None],
            [None, None, "Product", product],
            [None, None, "Serial Number", serial],
            [None, "Deployment Info", None, None],
            [None, None, "Deployment Number", str(deployment_number)],
            [None, None, "Logging Interval", f"0 hour {interval_minutes} minutes 0 seconds"],
        ):
            sheet.append(row)
        for name, unit, values in series:
            frame = pd.Series(values)
            sheet.append([f"Series : {name} , {unit}", None, None, None])
            sheet.append([None, "Series Statistics", None, None])
            sheet.append([None, None, "Samples", str(reported_samples or len(values))])
            sheet.append([None, None, "Max", f"{frame.max():.2f}"])
            sheet.append([None, None, "Min", f"{frame.min():.2f}"])
            sheet.append([None, None, "Avg", f"{frame.mean():.2f}"])

    book.save(path)
    return path


# --------------------------------------------------------------------------
# sniff
# --------------------------------------------------------------------------


def test_sniff_accepts_both_reference_files():
    assert hobo_xlsx.sniff(ORIGINAL) is True
    assert hobo_xlsx.sniff(EDITED) is True


def test_sniff_rejects_non_hobo_without_raising(tmp_path):
    """A file-drop directory holds whatever someone put there; sniff must not crash."""
    assert hobo_xlsx.sniff(REPO_ROOT / "README.md") is False
    assert hobo_xlsx.sniff(tmp_path / "does-not-exist.xlsx") is False

    not_a_workbook = tmp_path / "corrupt.xlsx"
    not_a_workbook.write_text("this is not a zip archive")
    assert hobo_xlsx.sniff(not_a_workbook) is False

    wrong_sheets = write_hobo_xlsx(tmp_path / "partial.xlsx", sheets=("Data",))
    assert hobo_xlsx.sniff(wrong_sheets) is False


# --------------------------------------------------------------------------
# parse -- the original export
# --------------------------------------------------------------------------


def test_parse_original_is_unedited():
    raw = hobo_xlsx.parse(ORIGINAL)
    assert raw.provenance == "original"
    assert raw.edit_signals == ()


def test_parse_original_reads_timezone_and_unit_from_headers():
    raw = hobo_xlsx.parse(ORIGINAL)
    assert raw.tz_token == "PDT"
    assert [(s.name, s.unit) for s in raw.series] == [("Tidbit 1", f"{DEGREE}F")]
    assert raw.series[0].column == SERIES_COLUMN


def test_parse_original_reproduces_pinned_statistics():
    """Same numbers as tests/test_fixtures_hobo.py, now via the adapter."""
    raw = hobo_xlsx.parse(ORIGINAL)
    info = raw.series[0]
    assert info.n == 3029
    assert len(raw.data) == 3029
    assert round(info.minimum, 2) == 58.60
    assert round(info.maximum, 2) == 75.35
    assert info.first == local(2026, 7, 11, 7, 0)
    assert info.last == local(2026, 8, 1, 7, 40)

    spacing = raw.data["timestamp_local"].diff().dropna().dt.total_seconds().unique()
    assert list(spacing) == [600.0]


def test_parse_leaves_conversion_to_the_normalizer():
    """Adapters extract faithfully and nothing more (docs/06 s4)."""
    raw = hobo_xlsx.parse(ORIGINAL)
    assert raw.data["timestamp_local"].dt.tz is None, "timestamps must stay tz-naive"
    assert raw.series[0].unit == f"{DEGREE}F", "degF must not be converted to degC"
    assert round(raw.series[0].minimum, 2) == 58.60, "the pre-window install transient is kept"


# --------------------------------------------------------------------------
# parse -- the hand-edited copy
# --------------------------------------------------------------------------


def test_parse_edited_is_flagged_with_every_structural_signal():
    raw = hobo_xlsx.parse(EDITED)
    assert raw.provenance == "edited"
    joined = " | ".join(raw.edit_signals)
    assert "Unnamed: 3" in joined  # hand-added helper column
    assert "without a valid Date-Time" in joined  # 6 blanked installation rows
    assert "formula cell" in joined  # =min(C:C) / =max(C:C)
    assert "contiguous run from 1" in joined  # counter starts at 7


def test_parse_edited_keeps_only_real_measurements():
    raw = hobo_xlsx.parse(EDITED)
    assert len(raw.data) == 3022
    assert raw.series[0].n == 3022
    assert round(raw.series[0].minimum, 2) == 63.96  # install transient gone
    assert raw.series[0].first == local(2026, 7, 11, 8, 0)
    assert raw.series[0].last == local(2026, 8, 1, 7, 30)


def test_parse_edited_ignores_the_helper_column():
    """`Unnamed: 3` is a signal, never a series."""
    raw = hobo_xlsx.parse(EDITED)
    assert [s.name for s in raw.series] == ["Tidbit 1"]
    assert set(raw.data["series_name"]) == {"Tidbit 1"}


def test_adapter_matches_the_hand_edit_row_for_row():
    """The adapter's edited-file read equals the original's registry window.

    docs/06 s3: applying the deployment window to the original reproduces the
    hand-edited file. Same claim as test_fixtures_hobo, but through the adapter.
    """
    original = hobo_xlsx.parse(ORIGINAL).data
    edited = hobo_xlsx.parse(EDITED).data
    deployment = find_deployment(load_registry(REGISTRY_PATH), KNOWN_SERIAL)
    start, end = deployment.window_local

    stamps = original["timestamp_local"]
    windowed = original.loc[(stamps >= start) & (stamps <= end)]
    assert windowed["timestamp_local"].tolist() == edited["timestamp_local"].tolist()
    assert windowed["value"].tolist() == edited["value"].tolist()


# --------------------------------------------------------------------------
# metadata
# --------------------------------------------------------------------------


def test_metadata_reads_device_deployment_and_events():
    meta = hobo_xlsx.metadata(ORIGINAL)
    assert meta["adapter"] == "hobo_xlsx"
    assert meta["serial"] == KNOWN_SERIAL
    assert meta["product"] == "MX2204"
    assert meta["deployment_number"] == 3
    assert meta["logging_interval_seconds"] == 600
    assert meta["provenance"] == "original"
    assert meta["filename"]["serial"] == KNOWN_SERIAL

    events = {(e["event"], e["timestamp"]) for e in meta["events"]}
    assert ("Started", "2026-07-11T07:00:00") in events
    assert ("End of File", "2026-08-01T07:40:00") in events
    assert sum(1 for e in meta["events"] if e["event"] == "Host Connected") == 3

    stats = meta["series"][0]["details_statistics"]
    assert stats["Samples"] == "3029"
    assert stats["Min"] == "58.60"
    assert stats["Std Dev"] == "2.38"  # recorded, though docs/06 s5.1 asserts only n/min/max/mean


def test_metadata_of_a_renamed_file_falls_back_to_details():
    meta = hobo_xlsx.metadata(EDITED)
    assert meta["filename"]["serial"] is None
    assert meta["serial"] == KNOWN_SERIAL


# --------------------------------------------------------------------------
# validation -- docs/06 s5
# --------------------------------------------------------------------------


def test_validate_original_passes_every_check(registry):
    report = hobo_xlsx.validate(ORIGINAL, registry=registry)
    assert report.provenance == "original"
    assert report.quarantined is False
    assert report.ok is True
    assert report.warnings == ()
    assert {c.name: c.status for c in report.checks} == {
        "details_statistics": "pass",
        "events_consistency": "pass",
        "cadence_audit": "pass",
        "timezone_crosscheck": "pass",
        "filename_serial": "pass",
        REGISTRY_GATE: "pass",
    }


def test_validate_edited_skips_consistency_checks_with_a_warning(registry):
    """docs/06 s3: skipped, never silently passed."""
    report = hobo_xlsx.validate(EDITED, registry=registry)
    assert report.provenance == "edited"
    assert report.check("details_statistics").status == "skipped"
    assert report.check("events_consistency").status == "skipped"
    assert any("details_statistics" in w for w in report.warnings)
    assert any("events_consistency" in w for w in report.warnings)

    # The audit and the gate still run on an edited file.
    assert report.check("cadence_audit").status == "pass"
    assert report.check(REGISTRY_GATE).status == "pass"
    assert report.quarantined is False


def test_statistics_check_still_fails_on_a_structurally_clean_file(tmp_path, registry):
    """Provenance is structural, so check 1 can still catch a truncated export.

    If `edited` were inferred from a statistics mismatch, this file would be
    excused as a hand edit and the cross-check could never fail at all.
    """
    path = write_hobo_xlsx(tmp_path / "truncated.xlsx", reported_samples=9999)
    report = hobo_xlsx.validate(path, registry=registry)
    assert report.provenance == "original"
    assert report.check("details_statistics").status == "fail"
    assert "Samples='9999'" in report.check("details_statistics").detail
    assert report.ok is False


def test_cadence_audit_reports_a_gap(tmp_path, registry):
    path = write_hobo_xlsx(
        tmp_path / "gappy.xlsx",
        series=(("Tidbit 1", f"{DEGREE}F", [60.0, 61.0, 62.0, 63.0]),),
        gap_after=1,
    )
    report = hobo_xlsx.validate(path, registry=registry)
    cadence = report.check("cadence_audit")
    assert cadence.status == "warn"
    assert "gap" in cadence.detail
    assert any("cadence_audit" in w for w in report.warnings)


# --------------------------------------------------------------------------
# The registry gate -- docs/06 s5 check 4
# --------------------------------------------------------------------------


def test_unregistered_serial_is_quarantined(tmp_path, registry):
    path = write_hobo_xlsx(tmp_path / "stranger.xlsx", serial="99999999")
    report = hobo_xlsx.validate(path, registry=registry)

    gate = report.check(REGISTRY_GATE)
    assert gate.status == "fail"
    assert "99999999" in gate.detail
    assert "quarantine" in gate.detail
    assert report.quarantined is True
    assert report.ok is False


def test_incomplete_deployment_record_is_quarantined(tmp_path):
    """A serial match is not enough: docs/06 requires tz, window, and series map."""
    thin = tmp_path / "sites.json"
    thin.write_text(
        '{"sites": [{"site_id": "PROJ:X", "deployments": '
        '[{"serial": "22506632", "deployment_number": 1}]}]}',
        encoding="utf-8",
    )
    report = hobo_xlsx.validate(ORIGINAL, registry=load_registry(thin))
    gate = report.check(REGISTRY_GATE)
    assert gate.status == "fail"
    assert "tz" in gate.detail
    assert "window_local" in gate.detail
    assert "series_map" in gate.detail
    assert report.quarantined is True


def test_missing_series_map_is_quarantined(tmp_path):
    """Timed and placed but unmapped: the normalizer would have to guess."""
    thin = tmp_path / "sites.json"
    thin.write_text(
        json.dumps(
            {
                "sites": [
                    {
                        "site_id": "PROJ:X",
                        "deployments": [
                            {
                                "serial": KNOWN_SERIAL,
                                "deployment_number": 1,
                                "tz": "America/Los_Angeles",
                                "window_local": ["2026-07-11 08:00", "2026-08-01 07:30"],
                            }
                        ],
                    }
                ]
            }
        ),
        encoding="utf-8",
    )
    report = hobo_xlsx.validate(ORIGINAL, registry=load_registry(thin))
    gate = report.check(REGISTRY_GATE)
    assert gate.status == "fail"
    assert gate.detail.endswith("-- quarantine")
    assert "missing series_map)" in gate.detail  # only the absent field is named
    assert report.quarantined is True


def test_series_map_resolves_the_parameter_by_series_name(registry):
    """The key is the header's sensor name, not the full `{name} , {unit}` label."""
    deployment = find_deployment(registry, KNOWN_SERIAL)
    assert deployment.parameter_for("Tidbit 1") == "sea_water_temperature"
    assert deployment.parameter_for(SERIES_COLUMN) is None
    assert deployment.parameter_for("Light") is None


def test_gate_passes_although_the_position_is_unverified(registry):
    """Serial 22506632's lat/lon are deliberately null pending a GPS fix."""
    site = registry.site("PROJ:YELLOW-BUOY")
    assert site["lat"] is None and site["lon"] is None

    deployment = find_deployment(registry, KNOWN_SERIAL)
    assert deployment.depth_m is None
    assert deployment.is_complete is True

    report = hobo_xlsx.validate(ORIGINAL, registry=registry)
    assert report.check(REGISTRY_GATE).status == "pass"
    assert report.quarantined is False


def test_deployment_cannot_carry_a_position():
    """Enforced by the type, not by a reviewer remembering the rule."""
    fields = Deployment.__dataclass_fields__
    assert "lat" not in fields and "lon" not in fields


# --------------------------------------------------------------------------
# Nothing is assumed: units, timezone, series count
# --------------------------------------------------------------------------


def test_unit_and_timezone_come_from_the_headers(tmp_path, registry):
    """HOBOconnect is configurable; a teammate's export may be degC/PST."""
    path = write_hobo_xlsx(
        tmp_path / "celsius.xlsx",
        tz="PST",
        series=(("Buoy Probe", f"{DEGREE}C", [14.0, 14.5, 15.0]),),
        start=local(2026, 1, 10, 7, 0),
    )
    raw = hobo_xlsx.parse(path)
    assert raw.tz_token == "PST"
    assert [(s.name, s.unit) for s in raw.series] == [("Buoy Probe", f"{DEGREE}C")]

    report = hobo_xlsx.validate(path, registry=registry)
    assert report.check("timezone_crosscheck").status == "pass"  # January is PST


def test_timezone_crosscheck_warns_when_the_token_disagrees(tmp_path, registry):
    path = write_hobo_xlsx(
        tmp_path / "wrong-token.xlsx",
        tz="PST",
        start=local(2026, 7, 11, 7, 0),  # July is PDT in America/Los_Angeles
    )
    check = hobo_xlsx.validate(path, registry=registry).check("timezone_crosscheck")
    assert check.status == "warn"
    assert "PST" in check.detail and "PDT" in check.detail


def test_multi_series_logger_yields_one_series_each(tmp_path, registry):
    """Other HOBO models export temperature plus light (docs/06 s6)."""
    path = write_hobo_xlsx(
        tmp_path / "two-series.xlsx",
        series=(
            ("Temp", f"{DEGREE}C", [14.0, 14.5, 15.0]),
            ("Light", "lux", [100.0, 220.0, 340.0]),
        ),
    )
    raw = hobo_xlsx.parse(path)
    assert [(s.name, s.unit) for s in raw.series] == [("Temp", f"{DEGREE}C"), ("Light", "lux")]
    assert len(raw.data) == 6
    assert raw.series_frame("Light")["value"].tolist() == [100.0, 220.0, 340.0]

    report = hobo_xlsx.validate(path, registry=registry)
    assert report.check("details_statistics").status == "pass"
