"""`kelpcompare ingest --source project`, end to end (docs/03, docs/06 s5).

Every test runs the real command against a `tmp_path` data root, so what is
asserted is what an operator would get. Nothing here touches the repo's own
`data/` -- raw is append-only forever (hard rule 1).

The two reference binaries are copied into a temporary drop directory rather
than ingested in place, which is also the honest simulation: the operator copies
an export off the instrument into `incoming/`.
"""

from __future__ import annotations

import json
import shutil
from pathlib import Path

import pytest
from click.testing import CliRunner

from kelpcompare import storage
from kelpcompare.cli import main
from kelpcompare.registry import find_deployment, load_registry
from kelpcompare.storage import Zones

REPO_ROOT = Path(__file__).resolve().parents[1]
FIX = Path(__file__).parent / "fixtures"
ORIGINAL = FIX / "Tidbit_1__22506632__2026-08-01_07_44_27_PDT__Data_PDT_.xlsx"
EDITED = FIX / "yellow_buoy_temps.xlsx"
REGISTRY_SOURCE = REPO_ROOT / "data" / "registry"

KNOWN_SERIAL = "22506632"

#: The site that serial is deployed at, read from the committed registry rather than
#: pinned here. Renaming a site is a data(registry) edit; it must not need a code
#: change to keep this green.
PROJECT_SITE = find_deployment(load_registry(REGISTRY_SOURCE / "sites.json"), KNOWN_SERIAL).site_id


@pytest.fixture
def data_root(tmp_path) -> Path:
    """A docs/03 data root with the committed registry and an empty incoming/."""
    root = tmp_path / "data"
    (root / "registry").mkdir(parents=True)
    for name in ("sites.json", "parameters.json"):
        shutil.copy2(REGISTRY_SOURCE / name, root / "registry" / name)
    (root / "raw" / "project_sensors" / "incoming").mkdir(parents=True)
    return root


def incoming(data_root: Path) -> Path:
    return data_root / "raw" / "project_sensors" / "incoming"


def drop(data_root: Path, fixture: Path, name: str | None = None) -> Path:
    target = incoming(data_root) / (name or fixture.name)
    shutil.copy2(fixture, target)
    return target


def run_ingest(data_root: Path, *extra: str):
    result = CliRunner().invoke(
        main, ["ingest", "--source", "project", "--data-root", str(data_root), *extra]
    )
    if result.exception and not isinstance(result.exception, SystemExit):
        raise result.exception
    return result


def manifest_of(data_root: Path) -> dict:
    files = sorted((data_root / "raw" / "_manifests").glob("*.json"))
    assert len(files) == 1, f"expected one manifest, found {[f.name for f in files]}"
    return json.loads(files[0].read_text(encoding="utf-8"))


# --------------------------------------------------------------------------
# The happy path
# --------------------------------------------------------------------------


def test_ingesting_the_reference_export_writes_the_documented_rows(data_root):
    drop(data_root, ORIGINAL)
    result = run_ingest(data_root)
    assert result.exit_code == 0

    stored = storage.read_observations(Zones.at(data_root), source="project")
    assert len(stored) == 3029
    assert set(stored["parameter"]) == {"sea_water_temperature"}
    assert set(stored["site_id"]) == {PROJECT_SITE}

    usable = stored.loc[stored["qc_flag"] <= 2]
    assert len(usable) == 3022
    assert round(usable["value"].min(), 2) == 17.76
    assert round(stored["value"].min(), 2) == 14.78  # the flagged install transient


def test_rows_land_in_the_documented_partition(data_root):
    drop(data_root, ORIGINAL)
    run_ingest(data_root)

    partition = data_root / "observations" / "source=project" / "year=2026"
    parts = list(partition.glob("part-*.parquet"))
    assert len(parts) == 1


def test_the_raw_landing_is_content_addressed_and_keeps_the_original(data_root):
    """Copy, never move: hard rule 1 forbids deleting from raw/."""
    dropped = drop(data_root, ORIGINAL)
    run_ingest(data_root)

    landed = list((data_root / "raw" / "project_sensors" / KNOWN_SERIAL).iterdir())
    assert len(landed) == 1
    assert landed[0].name.endswith(f"__{ORIGINAL.name}")
    assert landed[0].read_bytes() == ORIGINAL.read_bytes()
    assert dropped.exists()  # the operator clears incoming/, not the pipeline


def test_the_manifest_records_the_checks_and_the_landing(data_root):
    drop(data_root, ORIGINAL)
    run_ingest(data_root)
    payload = manifest_of(data_root)

    assert payload["command"] == "ingest"
    assert payload["sources"] == ["project"]
    assert payload["counts"] == {"ingested": 1}

    entry = payload["files"][0]
    assert entry["outcome"] == "ingested"
    assert entry["serial"] == KNOWN_SERIAL
    assert entry["site_id"] == PROJECT_SITE
    assert entry["provenance"] == "original"
    assert entry["rows_in"] == 3029
    assert entry["rows_out"] == 3029
    assert entry["qc_flags"] == {"2": 3022, "4": 7}
    assert entry["landed"].endswith(ORIGINAL.name)
    assert {c["name"] for c in entry["checks"]} == {
        "details_statistics",
        "events_consistency",
        "cadence_audit",
        "timezone_crosscheck",
        "filename_serial",
        "registry_gate",
        "series_mapping",
    }


def test_an_edited_file_is_ingested_and_marked(data_root):
    """docs/06 s3: accepted when it is all we have, but never silently."""
    drop(data_root, EDITED)
    assert run_ingest(data_root).exit_code == 0

    entry = manifest_of(data_root)["files"][0]
    assert entry["outcome"] == "ingested"
    assert entry["provenance"] == "edited"
    assert any("details_statistics" in w for w in entry["warnings"])
    assert any("events_consistency" in w for w in entry["warnings"])


# --------------------------------------------------------------------------
# Quarantine -- docs/06 s5 check 4, hard rule 5
# --------------------------------------------------------------------------


def test_an_unregistered_serial_is_quarantined_and_nothing_is_stored(data_root, tmp_path):
    stranger = _reserialed(tmp_path, ORIGINAL, "99999999")
    drop(data_root, stranger)
    result = run_ingest(data_root)
    assert result.exit_code == 0  # fail soft: a rejection is not a crash

    assert (data_root / "quarantine" / stranger.name).exists()
    assert not (data_root / "observations").exists()
    assert not (data_root / "raw" / "project_sensors" / "99999999").exists()

    entry = manifest_of(data_root)["files"][0]
    assert entry["outcome"] == "quarantined"
    assert entry["landed"] is None  # a quarantined file never enters raw/
    assert "99999999" in entry["reason"]


def test_a_quarantined_file_stays_in_incoming_for_a_retry(data_root, tmp_path):
    """Fix the registry, re-run, and the file is picked straight back up."""
    stranger = _reserialed(tmp_path, ORIGINAL, "99999999")
    dropped = drop(data_root, stranger)
    run_ingest(data_root)
    assert dropped.exists()


def test_one_bad_file_does_not_cost_the_run_the_good_one(data_root, tmp_path):
    drop(data_root, ORIGINAL)
    drop(data_root, _reserialed(tmp_path, ORIGINAL, "99999999"))
    run_ingest(data_root)

    assert manifest_of(data_root)["counts"] == {"ingested": 1, "quarantined": 1}
    assert len(storage.read_observations(Zones.at(data_root))) == 3029


def test_a_file_no_adapter_recognizes_is_skipped(data_root):
    (incoming(data_root) / "notes.txt").write_text("field notes", encoding="utf-8")
    drop(data_root, ORIGINAL)
    run_ingest(data_root)

    payload = manifest_of(data_root)
    assert payload["counts"] == {"ingested": 1, "skipped": 1}
    skipped = next(f for f in payload["files"] if f["outcome"] == "skipped")
    assert "no adapter" in skipped["reason"]


def test_a_deployment_missing_its_series_map_is_quarantined(data_root):
    """The gate now covers the mapping the normalizer needs (docs/06 s5 check 4)."""
    sites = data_root / "registry" / "sites.json"
    payload = json.loads(sites.read_text(encoding="utf-8"))
    record = next(
        d
        for site in payload["sites"]
        for d in site.get("deployments", ())
        if str(d.get("serial")) == KNOWN_SERIAL
    )
    record.pop("series_map")
    sites.write_text(json.dumps(payload), encoding="utf-8")

    drop(data_root, ORIGINAL)
    run_ingest(data_root)

    entry = manifest_of(data_root)["files"][0]
    assert entry["outcome"] == "quarantined"
    assert "series_map" in entry["reason"]


# --------------------------------------------------------------------------
# Re-ingest and overlap -- docs/06 s5 check 5
# --------------------------------------------------------------------------


def test_reingesting_the_same_file_does_not_duplicate_rows(data_root):
    drop(data_root, ORIGINAL)
    run_ingest(data_root)
    run_ingest(data_root)
    assert len(storage.read_observations(Zones.at(data_root))) == 3029


def test_overlapping_readouts_of_one_logger_dedupe(data_root):
    """The edited file's 3022 rows are a subset of the original's 3029."""
    drop(data_root, ORIGINAL)
    drop(data_root, EDITED)
    run_ingest(data_root)

    stored = storage.read_observations(Zones.at(data_root))
    assert len(stored) == 3029
    assert not stored.duplicated(subset=["site_id", "parameter", "timestamp"]).any()


# --------------------------------------------------------------------------
# Options
# --------------------------------------------------------------------------


def test_dry_run_writes_nothing_at_all(data_root):
    drop(data_root, ORIGINAL)
    result = run_ingest(data_root, "--dry-run")

    assert result.exit_code == 0
    assert "dry run" in result.output
    assert not (data_root / "observations").exists()
    assert not (data_root / "raw" / "_manifests").exists()
    assert not (data_root / "raw" / "project_sensors" / KNOWN_SERIAL).exists()


def test_a_single_file_can_be_named_directly(data_root, tmp_path):
    elsewhere = tmp_path / "off-tree" / ORIGINAL.name
    elsewhere.parent.mkdir()
    shutil.copy2(ORIGINAL, elsewhere)

    result = run_ingest(data_root, "--path", str(elsewhere))
    assert result.exit_code == 0
    assert len(storage.read_observations(Zones.at(data_root))) == 3029


def test_an_empty_drop_directory_is_not_an_error(data_root):
    result = run_ingest(data_root)
    assert result.exit_code == 0
    assert "nothing to ingest" in result.output
    assert not (data_root / "raw" / "_manifests").exists()


def test_an_unimplemented_source_says_so(data_root):
    """`cdip` is still only a docs/02 entry; `ndbc` stopped being one."""
    result = CliRunner().invoke(main, ["ingest", "--source", "cdip", "--data-root", str(data_root)])
    assert result.exit_code != 0
    assert "not implemented" in str(result.exception)
    # The message lists what an operator can actually run, both shapes together.
    assert "ndbc" in str(result.exception)
    assert "project" in str(result.exception)


# --------------------------------------------------------------------------
# Building an unregistered file from the reference export
# --------------------------------------------------------------------------


def _reserialed(tmp_path: Path, source: Path, serial: str) -> Path:
    """The reference export with a serial the registry has never heard of."""
    import openpyxl

    target = tmp_path / f"stranger_{serial}.xlsx"
    book = openpyxl.load_workbook(source)
    details = book["Details"]
    for row in details.iter_rows():
        for cell in row:
            if str(cell.value).strip() == "Serial Number":
                details.cell(row=cell.row, column=cell.column + 1, value=serial)
    book.save(target)
    book.close()
    return target
