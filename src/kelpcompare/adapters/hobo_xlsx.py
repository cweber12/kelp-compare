"""The `hobo_xlsx` adapter: `sniff`/`parse`/`metadata` plus the workbook loader.

This module owns the *container* -- opening an xlsx and getting three sheets out
of it. What the sheets mean lives in `hobo_common`, so `hobo_csv` can be added
later as a different loader over identical semantics (docs/06 s6).

Nothing here converts anything. Timezone, units and the deployment-window trim
are the normalizer's job; quarantining a file is the ingest CLI's (docs/06 s4).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from kelpcompare.adapters import hobo_common
from kelpcompare.adapters.base import RawSeries, ValidationReport
from kelpcompare.adapters.hobo_common import (
    DATA_SHEET,
    DETAILS_SHEET,
    EVENTS_SHEET,
    REQUIRED_SHEETS,
    HoboSheets,
)
from kelpcompare.registry import Registry, load_registry

ADAPTER_NAME = "hobo_xlsx"
SUFFIXES = (".xlsx", ".xlsm")


def sniff(path: Path | str) -> bool:
    """Cheap check: is this a HOBOconnect workbook?

    Never raises. A file-drop directory contains whatever someone put there, and
    a malformed file must fall through to the next adapter, not crash the run.
    """
    candidate = Path(path)
    if candidate.suffix.lower() not in SUFFIXES:
        return False
    try:
        book = openpyxl.load_workbook(candidate, read_only=True)
        try:
            names = set(book.sheetnames)
        finally:
            book.close()
    except Exception:  # noqa: BLE001 -- "can I parse this?" has one honest answer on failure
        return False
    return set(REQUIRED_SHEETS).issubset(names)


def parse(path: Path | str) -> RawSeries:
    """Measurements plus per-series metadata, extracted verbatim (docs/06 s4)."""
    return hobo_common.build_raw_series(_load(path))


def metadata(path: Path | str) -> dict:
    """Serial, model, interval, events, and export-time statistics (docs/06 s4)."""
    sheets = _load(path)
    payload = hobo_common.read_metadata(sheets, hobo_common.build_raw_series(sheets))
    payload["adapter"] = ADAPTER_NAME
    return payload


def validate(path: Path | str, registry: Registry | None = None) -> ValidationReport:
    """Run the docs/06 s5 checks. Returns a verdict; takes no action.

    `report.quarantined` means the registry gate failed. Moving the file into
    `data/quarantine/` is the ingest CLI's job (docs/03) -- one place decides
    what happens to a file.
    """
    sheets = _load(path)
    raw = hobo_common.build_raw_series(sheets)
    resolved = registry if registry is not None else load_registry()
    return ValidationReport(
        path=sheets.path,
        provenance=raw.provenance,
        checks=hobo_common.run_checks(sheets, raw, resolved),
    )


def _load(path: Path | str) -> HoboSheets:
    """Read the three sheets. The only part of the adapter that knows about xlsx."""
    source = Path(path)
    with pd.ExcelFile(source) as book:
        missing = set(REQUIRED_SHEETS) - set(book.sheet_names)
        if missing:
            raise ValueError(
                f"{source}: not a HOBOconnect export, missing sheet(s) {sorted(missing)}"
            )
        data = book.parse(DATA_SHEET)
        events = book.parse(EVENTS_SHEET)
        # Details has no header row -- it is a section/group/key/value hierarchy.
        details = book.parse(DETAILS_SHEET, header=None)

    return HoboSheets(
        path=source,
        data=data,
        events=events,
        details=details,
        formula_cells=_formula_cells(source),
    )


def _formula_cells(path: Path) -> tuple[str, ...]:
    """Coordinates of formula cells in `Data` -- a hand-edit signal a values-only
    read cannot see (the reference edited file carries `=min(C:C)` in D2/D3)."""
    book = openpyxl.load_workbook(path, data_only=False)
    try:
        if DATA_SHEET not in book.sheetnames:
            return ()
        sheet = book[DATA_SHEET]
        return tuple(
            cell.coordinate for row in sheet.iter_rows() for cell in row if cell.data_type == "f"
        )
    finally:
        book.close()
